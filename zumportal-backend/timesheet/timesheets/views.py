from datetime import datetime
from rest_framework.decorators import APIView
from rest_framework.generics import ListAPIView

from authentication.models import User 
from project.models import Project
from .models import Timesheet,Task
from .serializers import TimesheetSerializer,AddTimesheetSerialzer
from rest_framework.response import Response
from rest_framework import status
import csv
from django.http import HttpResponse
from openpyxl import Workbook
from io import TextIOWrapper
import openpyxl
from django.db.models import Sum
from django.conf import settings
from django.db.models import Q
from rest_framework.permissions import IsAuthenticated
from .permissions import CanAddTimesheet,CanChangeTimesheet,CanDeleteTimesheet,CanExportTimesheet,CanImportTimesheet,CanViewTimesheet


class EmployeeTimesheetsAPI(APIView):
    permission_classes=[IsAuthenticated]
    def get(self, request, format=None):
        employees = User.objects.filter(role__in=[1,2,3])

        month = request.GET.get('month')  # Get the month query parameter
        
        data = []
        for employee in employees:
            profile_photo_url = None
            
            # Check if the employee has a profile photo
            if employee.profile_photo:
                profile_photo_url = request.build_absolute_uri(settings.MEDIA_URL + str(employee.profile_photo))

            projects = Project.objects.filter(Q(assigned_to=employee) | Q(project_manager=employee)| Q(scrum_master=employee)).distinct()
            projects_data = []
            total_billable_hours = 0
            total_not_billable_hours = 0
            timesheets_data = []

            # Get timesheets for the employee
            timesheets = Timesheet.objects.filter(employee=employee )
            
            # Filter timesheets by the specified month
            if month:
                timesheets = timesheets.filter(date__month=month)

            # Collect timesheet data and calculate the sum of hours for each month
            monthly_billable_hours = {}
            monthly_not_billable_hours = {}
            for timesheet in timesheets:
                date_month = timesheet.date.strftime('%Y-%m')
                billable_hours = timesheet.billablehours
                not_billable_hours = timesheet.notbillablehours
                
                # Update monthly totals
                monthly_billable_hours[date_month] = monthly_billable_hours.get(date_month, 0) + billable_hours
                monthly_not_billable_hours[date_month] = monthly_not_billable_hours.get(date_month, 0) + not_billable_hours

                # Collect timesheet data
                timesheets_data.append({
                    'date': timesheet.date,
                    'billablehours': billable_hours,
                    'notbillablehours': not_billable_hours,
                    # Add other fields as needed
                })

            # Collect project data and aggregate hours across all projects for the specified month
            for project in projects:                
                project_billable_hours = timesheets.filter(project=project).aggregate(Sum('billablehours'))['billablehours__sum'] or 0
                project_not_billable_hours = timesheets.filter(project=project).aggregate(Sum('notbillablehours'))['notbillablehours__sum'] or 0

                projects_data.append({
                    'id': project.id,
                    'name': project.name,
                    'abbreviation': project.abbreviation,
                    'starter_at': project.starter_at,
                    'end_date': project.end_date,
                    'total_billable_hours': project_billable_hours,
                    'total_not_billable_hours': project_not_billable_hours
                })
                
                total_billable_hours += project_billable_hours
                total_not_billable_hours += project_not_billable_hours

            data.append({
                'employee': {
                    'id': employee.id,
                    'firstname': employee.firstname,
                    'lastname': employee.lastname,
                    'profile_photo': profile_photo_url,
                    'position': employee.position
                },
                'projects': projects_data,
                'total_billable_hours': total_billable_hours,
                'total_not_billable_hours': total_not_billable_hours,
                'timesheets': timesheets_data,
                'monthly_billable_hours': monthly_billable_hours,
                'monthly_not_billable_hours': monthly_not_billable_hours
            })
        return Response(data, status=status.HTTP_200_OK)

class AllTimesheetsAPI(APIView):
    permission_classes=[IsAuthenticated]
    def get(self, request, format=None):
        timesheets = Timesheet.objects.all()
        serializer = TimesheetSerializer(timesheets, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class TimesheetAPI(APIView):
    permission_classes=[IsAuthenticated]
    def get(self, request, format=None, *args, **kwargs):       
        timesheet_id = self.kwargs.get('id')
        print(timesheet_id)

        if timesheet_id:
            try:
                timesheet = Timesheet.objects.get(id=timesheet_id)

                timesheet_serializer = TimesheetSerializer(timesheet, context={'request': request})
                
                data = timesheet_serializer.data

                return Response(data, status=status.HTTP_200_OK)
            except Timesheet.DoesNotExist:
                return Response({'error': 'Timesheet not found'}, status=status.HTTP_404_NOT_FOUND)

        else:
            print("here all timesheets")
            categories = Timesheet.objects.all()
            serializer = TimesheetSerializer(categories, many=True,context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        
    def post(self, request, format=None, *args):
        self.permission_classes = [CanAddTimesheet]
        self.check_permissions(request)
        serializer = AddTimesheetSerialzer(data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
    def put(self, request, id, format=None):
        self.permission_classes = [CanChangeTimesheet]
        self.check_permissions(request)
        try:
            timesheet = Timesheet.objects.get(id=id)
            serializer = AddTimesheetSerialzer(instance=timesheet, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Timesheet.DoesNotExist:
            return Response({'error': 'Timesheet not found'}, status=status.HTTP_404_NOT_FOUND)
    
    def delete(self, request, format=None, **kwargs):
        self.permission_classes = [CanDeleteTimesheet]
        self.check_permissions(request)
        timesheet_id = kwargs.get('id')  # Use kwargs to get the 'id' parameter
        
        try:
            timesheet = Timesheet.objects.get(id=timesheet_id)
            tasks_deleted = timesheet.tasks.all().delete()  # Delete tasks associated with the timesheet

            timesheet.delete()  # Delete the timesheet
            return Response({'message': 'Timesheet and related tasks deleted successfully'}, status=status.HTTP_204_NO_CONTENT)
        except Timesheet.DoesNotExist:
            return Response({'error': 'Timesheet not found'}, status=status.HTTP_404_NOT_FOUND)
        except:
            return Response({'error': 'Bara laweg aalech'}, status=status.HTTP_404_NOT_FOUND)
class TimesheetListAPIView(ListAPIView):
    permission_classes=[IsAuthenticated]
    serializer_class = TimesheetSerializer

    def get_queryset(self):
        project_id = self.kwargs.get('project_id')
        if project_id:
            return Timesheet.objects.filter(project_id=project_id).order_by('-date', 'employee_id')

    
        
class CSVAPI(APIView):
    permission_classes=[IsAuthenticated]
    def get(self, request, format=None):
        try:
            timesheets = Timesheet.objects.all()
            csv_data = self.serialize_to_csv(timesheets)

            # Create an HTTP response with CSV content
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="timesheets.xlsx"'

            # Write the CSV data to the response
            writer = csv.writer(response)
            writer.writerows(csv_data)

            return response
        except Timesheet.DoesNotExist:
            return Response({'error': 'No timesheets found'}, status=status.HTTP_404_NOT_FOUND)

    def serialize_to_csv(self, timesheets):
        csv_data = [['Date', 'Site', 'Location', 'Hours', 'Billable', 'Task', 'Status']]

        for timesheet in timesheets:
            serializer = TimesheetSerializer(instance=timesheet)
            tasks = timesheet.tasks.all().values_list('task', 'status')

            # Add timesheet data
            csv_data.append([serializer.data['date'], serializer.data['site'], serializer.data['location'],
                             serializer.data['hours'], serializer.data['billable'], '', ''])

            # Add tasks data
            for task in tasks:
                csv_data.append(['', '', '', '', '', task[0], task[1]])

        return csv_data
    
class ExcelAPI(APIView):
    permission_classes=[IsAuthenticated]
    def get(self, request, year=None , month=None, employee_id=None , format=None):
        try:
            timesheets = Timesheet.objects.all()

            if employee_id:
                timesheets = timesheets.filter(employee_id=employee_id)
            if month:
                month = int(month)
                if month < 1 or month > 12:
                    return Response({'error': 'Invalid month'}, status=status.HTTP_400_BAD_REQUEST)
                timesheets = timesheets.filter(date__month=month)

            if year:
                timesheets = timesheets.filter(date__year=year)

            timesheets = timesheets.order_by('date')
            workbook = self.serialize_to_excel(timesheets)

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="timesheets.xlsx"'

            workbook.save(response)

            return response
        except Timesheet.DoesNotExist:
            return Response({'error': 'No timesheets found'}, status=status.HTTP_404_NOT_FOUND)

    def serialize_to_excel(self, timesheets):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['Employee Name', 'Employee Last Name', 'Employee Email', 'Date', 'Site', 'Location', 'Billable Hours', 'Not Billable Hours', 'Task', 'Status', 'Billable'])

        total_billable_hours = 0
        total_not_billable_hours = 0

        for timesheet in timesheets:
            serializer = TimesheetSerializer(instance=timesheet)
            tasks = timesheet.tasks.all().values_list('task', 'status' , 'billable')

            employee_name = timesheet.employee.firstname
            employee_last_name = timesheet.employee.lastname
            employee_email = timesheet.employee.email

            billable_hours = serializer.data['billablehours']
            not_billable_hours = serializer.data['notbillablehours']

            total_billable_hours += billable_hours
            total_not_billable_hours += not_billable_hours

            worksheet.append([
                employee_name,
                employee_last_name,
                employee_email,
                serializer.data['date'], 
                serializer.data['site'], 
                serializer.data['location'],
                billable_hours, 
                not_billable_hours, 
                '', 
                '',
                ''
            ])

            for task in tasks:
                worksheet.append(['', '', '', '', '', '', '', '', task[0], task[1] , task[2]])

        # Add total billable hours and total not billable hours
        last_row_index = len(timesheets) + 1
        worksheet.append(['', '', '', '', '', '', '', '', '', '', ''])
        worksheet.append(['', '', '', '', '', '', '', 'Total Billable Hours:', total_billable_hours,  '', ''])
        worksheet.append(['', '', '', '', '', '', '' ,'Total Not Billable Hours:', total_not_billable_hours, '', ''])

        # Apply bold font to the total rows
        # for row in worksheet.iter_rows(min_row=last_row_index + 1, max_row=last_row_index + 3):
        #     for cell in row:
        #         cell.font = Font(bold=True)
        # Adjust column widths to fit content
        # for col in worksheet.columns:
        #     max_length = 0
        #     column = col[0].column_letter
        #     for cell in col:
        #         try:
        #             if len(str(cell.value)) > max_length:
        #                 max_length = len(cell.value)
        #         except:
        #             pass
        #     adjusted_width = (max_length + 2) * 1.2
        #     worksheet.column_dimensions[column].width = adjusted_width
        return workbook
    def post(self, request, format=None):
        if 'file' not in request.FILES:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']

        if file.name.endswith('.xlsx'):
            return self.import_from_excel(file)
        elif file.name.endswith('.csv'):
            return self.import_from_csv(file)
        else:
            return Response({'error': 'Invalid file format'}, status=status.HTTP_400_BAD_REQUEST)

    def import_from_excel(self, file):
        try:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            
            timesheet_data = {}  # Dictionary to store timesheet data
            
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Check if the row contains timesheet data
                if all(cell is not None for cell in row[:7]):
                    # Unpack values from row, handle empty cells appropriately
                    employee_id, project_id, date, site, location, billablehours,notbillablehours = row[:7]
                    if isinstance(date, datetime):  # Check if date is already datetime object
                        pass  # No need to convert
                    elif isinstance(date, str):  # Check if date is string
                        date = datetime.strptime(date, '%d/%m/%Y')
                    else:
                        raise ValueError("Invalid date format")

                    # Add timesheet data to dictionary
                    timesheet_data[(employee_id, project_id, date, site, location, billablehours,notbillablehours)] = []
                
                # Check if the row contains task data
                elif any(cell is not None for cell in row[7:]):
                    if not timesheet_data:
                        raise ValueError("Task found without preceding timesheet data")
                    
                    # Extract task, status, and billable data
                    task, Status, billable = row[7:10]  # Correct indexing
                    print(f"Row {row_index}: Task: {task}, Status: {Status}, Billable: {billable}")
                    
                    # Append task, status, and billable to the last added timesheet entry
                    timesheet_data[list(timesheet_data.keys())[-1]].append((task, Status, billable))
            
            # Create Timesheet and Task objects
            for timesheet_info, tasks_info in timesheet_data.items():
                print(timesheet_info)
                print(tasks_info)
                timesheet = Timesheet.objects.create(
                    employee_id=timesheet_info[0] if timesheet_info[0] else None,
                    project_id=timesheet_info[1] if timesheet_info[1] else None,
                    date=timesheet_info[2],
                    site=timesheet_info[3],
                    location=timesheet_info[4],
                    billablehours=timesheet_info[5],
                    notbillablehours=timesheet_info[6]
                )
                for task_info in tasks_info:
                    Task.objects.create(
                        timesheet=timesheet,
                        task=task_info[0],
                        status=task_info[1],
                        billable=task_info[2] if task_info[2] is not None else None  # Use 'billable' instead of 'Billable'
                    )
            
            print("Data imported successfully")
            return Response({'message': 'Data imported successfully'}, status=status.HTTP_201_CREATED)
        
        except Exception as e:
            print(f"Error importing data at row {row_index}: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)




    def import_from_csv(self, file):
        try:
            decoded_file = file.read().decode('utf-8-sig')
            csv_reader = csv.DictReader(TextIOWrapper(decoded_file, 'utf-8-sig'))

            timesheet_data = {}
            current_timesheet = None

            for row in csv_reader:
                # Check if the row contains timesheet data
                if all(row.values()):
                    # Extract timesheet data
                    employee_id = row['Employee ID']
                    project_id = row['Project ID']
                    date = row['Date'],
                    site = row['Site']
                    location = row['Location']
                    hours = row['Hours']
                    billable = row['Billable']
                    
                    # Create a new Timesheet object
                    current_timesheet = Timesheet.objects.create(
                        employee_id=employee_id,
                        project_id=project_id,
                        date=date,
                        site=site,
                        location=location,
                        hours=hours,
                        billable=billable
                    )
                    
                    # Initialize an empty list for tasks associated with this timesheet
                    timesheet_data[current_timesheet] = []
                else:
                    # If the row contains task data, append it to the list of tasks for the current timesheet
                    timesheet_data[current_timesheet].append((row['Task'], row['Status']))

            # Create Task objects associated with the corresponding Timesheet
            for timesheet, tasks_info in timesheet_data.items():
                for task_info in tasks_info:
                    Task.objects.create(
                        timesheet=timesheet,
                        task=task_info[0],
                        status=task_info[1]
                    )

            return Response({'message': 'Data imported successfully'}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)